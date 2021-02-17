# CoinTaxman
# Copyright (C) 2021  Carsten Docktor <https://github.com/provinzio>

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.

# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import csv
import datetime
import logging
from pathlib import Path
import re
from typing import Optional

import openpyxl

import config
import misc
import transaction as tr

log = logging.getLogger(__name__)


class Book:

    def __init__(self) -> None:
        self.operations: list[tr.Operation] = []

    def __bool__(self) -> bool:
        return bool(self.operations)

    def _read_binance(self, file_path: Path) -> None:
        platform = "binance"
        operation_mapping = {
            "Distribution": "Airdrop",
            "Savings Interest": "CoinLendInterest",
            "Savings purchase": "CoinLend",
            "Savings Principal redemption": "CoinLendEnd",
            "Commission History": "Commission",
        }

        with open(file_path, encoding="utf8") as f:
            reader = csv.reader(f)

            # Skip header.
            next(reader)

            for row, (_utc_time, account, operation, coin, _change, remark) in enumerate(reader, 2):
                # Parse data.
                utc_time = datetime.datetime.strptime(
                    _utc_time, "%Y-%m-%d %H:%M:%S")
                change = float(_change)
                operation = operation_mapping.get(operation, operation)
                if operation in (
                    "The Easiest Way to Trade",
                    "Small assets exchange BNB",
                    "Transaction Related",
                ):
                    operation = "Sell" if change < 0 else "Buy"
                change = abs(change)

                # Validate data.
                assert account == "Spot", "Other types than Spot are currently not supported. Please create an Issue or PR."
                assert operation
                assert coin
                assert change

                # Check for problems.
                if remark:
                    log.warning(
                        "I may missed a remark in %s:%i: `%s`.", file_path, row, remark)

                # Append operation to the correct list.
                try:
                    Op = getattr(tr, operation)
                except AttributeError:
                    log.warning(
                        "Could not recognize operation `%s` in binance file `%s:%i`.", operation, file_path, row)
                    continue

                o = Op(utc_time, platform, change, coin)
                self.operations.append(o)

    def _read_etoro(self, file_path: Path) -> None:
        platform = "etoro"
        etoro_fiat = "USD"

        wb = openpyxl.load_workbook(filename=file_path, read_only=True)
        ws = wb["Transactions Report"]

        deposit_regex = re.compile("(\d+(?:\.\d+)?) (\w+) (\w+)")
        for row, (_utc_time, _account_balance, operation, details, _position_id, _amount, _realized_equity_change, _realized_equity, nwa) in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Parse data.
            utc_time = datetime.datetime.strptime(
                _utc_time, "%Y-%m-%d %H:%M:%S")
            account_balance = float(_account_balance)
            position_id = misc.xint(_position_id)
            amount = float(_amount)
            realized_equity_change = float(_realized_equity_change)
            realized_equity = float(_realized_equity)

            # Validate data.
            assert amount

            # Append operation to the correct list.
            if operation == "Deposit":
                if m := deposit_regex.match(details):
                    _deposited_fund, coin, description = m.groups()
                    deposited_fund = float(_deposited_fund)
                    if coin == etoro_fiat:
                        deposit = tr.Deposit(utc_time, platform, amount, coin)
                        self.operations.append(deposit)
                    else:
                        # TODO Can we deduct the changing fees from tax?
                        sell = tr.Sell(utc_time, platform,
                                       deposited_fund, coin)
                        self.operations.append(sell)
                        buy = tr.Buy(utc_time, platform, amount, coin)
                        self.operations.append(buy)
                else:
                    log.warning(
                        "Unable to parse deposit detail. Ignoring the entry in %s:%i", file_path, row)

            elif operation == "Open Position":
                buy_coin, sell_coin = details.split("/")
                assert sell_coin == etoro_fiat

                # TODO Determine fee from spread
                # TODO Determine bought coins
                #      We require an eToro API Key for that.

                sell = tr.Sell(utc_time, platform, amount, sell_coin)
                self.operations.append(sell)
                buy = tr.Buy(utc_time, platform, buy_change, buy_coin)
                self.operations.append(buy)

            else:
                log.warning(
                    "Could not recognize operation `%s` in etoro file `%s:%i`.", operation, file_path, row)

        wb.close()

    def detect_exchange(self, file_path: Path) -> Optional[str]:
        if file_path.suffix == ".csv":
            with open(file_path, encoding="utf8") as f:
                reader = csv.reader(f)
                header = next(reader, None)

            expected_headers = {
                "binance": ['UTC_Time', 'Account',
                            'Operation', 'Coin', 'Change', 'Remark']
            }
            for exchange, expected in expected_headers.items():
                if header == expected:
                    return exchange

        elif file_path.suffix == ".xlsx":
            wb = openpyxl.load_workbook(filename=file_path, read_only=True)
            sheets = wb.sheetnames
            sheet_headers = {
                sheet: next(wb[sheet].iter_rows(values_only=True), None)
                for sheet in sheets
            }
            wb.close()

            expected_sheet_headers = {
                "etoro": {
                    'Account Details': ('Details', None),
                    'Closed Positions': ('Position ID', 'Action', 'Copy Trader Name', 'Amount', 'Units', 'Open Rate', 'Close Rate', 'Spread', 'Profit', 'Open Date', 'Close Date', 'Take Profit Rate', 'Stop Loss Rate', 'Rollover Fees And Dividends', 'Is Real', 'Leverage', 'Notes'),
                    'Transactions Report': ('Date', 'Account Balance', 'Type', 'Details', 'Position ID', 'Amount', 'Realized Equity Change', 'Realized Equity', 'NWA'),
                    'Financial Summary': ('Figure', 'Amount in USD', 'Tax Rate'),
                },
            }
            for exchange, expected in expected_sheet_headers.items():
                if sheet_headers == expected:
                    return exchange
        return None

    def read_file(self, file_path: Path) -> None:
        """Import transactions form an account statement.

        Detect the exchange of the file. The file will be ignored with a
        warning, if the detecting or reading functionality is not implemented.

        Args:
            file_path (Path): Path to account statment.
        """
        assert file_path.is_file()

        if exchange := self.detect_exchange(file_path):
            try:
                read_file = getattr(self, f"_read_{exchange}")
            except AttributeError:
                log.warning(
                    "Unable to read files from the exchange `%s`. Skipping `%s`.", exchange, file_path)
                return

            log.info("Reading file from exchange %s at %s",
                     exchange, file_path)
            read_file(file_path)
        else:
            log.warning(
                f"Unable to detect the exchange of file `{file_path}`. Skipping file.")

    def get_account_statement_paths(self, statements_dir: Path) -> list[Path]:
        """Return file paths of all account statements in `statements_dir`.

        Args:
            statements_dir (str): Folder in which account statements will be searched.

        Returns:
            list[Path]: List of account statement file paths.
        """
        file_paths: list[Path] = []

        if statements_dir.is_dir():
            for file_path in statements_dir.iterdir():
                # Ignore .gitkeep and temporary exel files.
                filename = file_path.stem
                if filename == ".gitkeep" or filename.startswith("~$"):
                    continue

                file_paths.append(file_path)
        return file_paths

    def read_files(self) -> bool:
        """Read all account statements from the folder specified in the config.

        Returns:
            bool: Return True if everything went as expected.
        """
        paths = self.get_account_statement_paths(config.ACCOUNT_STATMENTS_PATH)

        if not paths:
            log.warning("No account statement files located in %s.",
                        config.ACCOUNT_STATMENTS_PATH)
            return False

        for file_path in paths:
            self.read_file(file_path)

        if not bool(self):
            log.warning("Unable to import any data.")
            return False

        return True
